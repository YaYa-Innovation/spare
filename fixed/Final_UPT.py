from flask import Flask, jsonify, render_template ,request
from gpiozero import CPUTemperature
import threading
from threading import Thread
from datetime import datetime
import time
from pymodbus.constants import Endian
from pymodbus.payload import BinaryPayloadDecoder
from pymodbus.client import ModbusSerialClient
from pymodbus.client import ModbusTcpClient
from openpyxl import *
import os.path
from pathlib import Path
import smtplib

SERVER = "172.16.0.8"

path_to_file =("Furnace_UPT/rpiWebServer/Furnace_UPT.xlsx")
path = Path(path_to_file)
global kwh_data
global unit
#global final_format_date

kwh_data = 0
unit ="KWH"
access=False

if path.is_file():
    print("The file ")

else:
    workbook = Workbook()
    sheet = workbook.active
#    sheet.protection.sheet = True
#    sheet.protection.password = '433'
#    sheet.protection.enable()
    workbook.save("Furnace_UPT/rpiWebServer/Furnace_UPT.xlsx")
    print("The File Not In This Directory now The File Created")

#http://172.16.5.175/
app = Flask(__name__)

@app.route("/")
def home():

    return render_template("Furnace_UPT_User.html")
@app.route("/edit")
def edit():
    return render_template("Furnace_UPT.html")
@app.route("/admin")
def admin():
    return render_template("admin.html")
def full_night_mail():
	global current_date

	workbook = load_workbook(path_to_file,data_only=True)
	sheet = workbook.active
	r=sheet.max_row
	for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
	        for cell in row:
	                if cell.value ==(current_date):
	                        re=sheet.cell(row=cell.row, column=cell.column)
	                        print("content Row: ",cell.row)
	                        rr=cell.row
	match_date=rr
	print(match_date)
	#Full Night Energy Consumption Unit
	full_night_unit=sheet.cell(row = match_date,column=12).value
	full_night_charge_metal=sheet.cell(row = match_date,column=5).value
	full_night_upt=sheet.cell(row = match_date,column=15).value
	full_night_emp=sheet.cell(row = match_date,column=9).value

	FROM = "RPI@texmo.net"
	TO = ["ahs@texmo.net","nta@texmo.net","psu@texmo.net","svu@texmo.net","bks@texmo.net"]
	SUBJECT = "TEXMO INDUSTRIES FURNACE UPT "+current_date
	DATA = str(full_night_unit)
	DATA1= str(full_night_charge_metal)
	DATA2= str(full_night_upt)
	DATA3=str(full_night_emp)

	TEXT ="\n"+"DATE : "+current_date +"\n"+ "FULL NIGHT UPT DETAILES:"+"\n"+"UNIT CONSUMPTION : "+ DATA +"\n"+"CHARGED METAL IN TON : " +DATA1+"\n"+"UNITS PER TON : "+DATA2+"\n"+"EMPLOYE NAME :"+DATA3

	message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

	%s

	""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	server = smtplib.SMTP(SERVER)
	server.sendmail(FROM, TO, message)
	print ("FULL NIGHT MAIL SUCCESS FULLY SENDED")
	server.quit()

def day_shift_mail():
	global current_date

	workbook = load_workbook(path_to_file,data_only=True)
	sheet = workbook.active
	r=sheet.max_row
	for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
	        for cell in row:
	                if cell.value ==(current_date):
	                        re=sheet.cell(row=cell.row, column=cell.column)
	                        print("content Row: ",cell.row)
	                        rr=cell.row
	match_date=rr
	print(match_date)
        #DAY Energy Consumption Unit
	day_shift_unit=sheet.cell(row = match_date,column=13).value
	day_shift_charge_metal=sheet.cell(row = match_date,column=6).value
	day_shift_upt=sheet.cell(row = match_date,column=16).value
	day_shift_emp=sheet.cell(row = match_date,column=10).value



	FROM = "RPI@texmo.net"
	TO = ["ahs@texmo.net","nta@texmo.net","psu@texmo.net","svu@texmo.net","bks@texmo.net"]
#	TO = ["ahs@texmo.net"] #,"nta@texmo.net","svu@texmo.net"]
	SUBJECT = "TEXMO INDUSTRIES FURNACE UPT "+current_date
	DATA = str(day_shift_unit)
	DATA1= str(day_shift_charge_metal)
	DATA2= str(day_shift_upt)
	DATA3= str(day_shift_emp)

#	TEXT ="\n"+"DATE : "+current_date +"\n"+ "FULL NIGHT UPT DETAILES     :"+"\n"+"     UNIT CONSUMPTION : "+ DATA +"\n"+"     CHARGED METAL IN TON : " +DATA1+"\n"+"     UNITS PER TON         : "+DATA2
	TEXT ="\n"+"DATE : "+current_date +"\n"+ "DAY SHIFT UPT DETAILES:"+"\n"+"UNIT CONSUMPTION : "+ DATA +"\n"+"CHARGED METAL IN TON : " +DATA1+"\n"+"UNITS PER TON : "+DATA2+"\n"+"EMPLOYE NAME :"+DATA3


	message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

	%s

	""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	server = smtplib.SMTP(SERVER)
	server.sendmail(FROM, TO, message)
	print ("DAY SHIFT MAIL SUCCESS FULLY SENDED")
	server.quit()

def half_night_mail():
	global current_date

	workbook = load_workbook(path_to_file,data_only=True)
	sheet = workbook.active
	r=sheet.max_row
	for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
	        for cell in row:
	                if cell.value ==(current_date):
	                        re=sheet.cell(row=cell.row, column=cell.column)
	                        print("content Row: ",cell.row)
	                        rr=cell.row
	match_date=rr
	print(match_date)

	#Full Night Energy Consumption Unit
	half_night_unit=sheet.cell(row = match_date,column=14).value
	half_night_charge_metal=sheet.cell(row = match_date,column=7).value
	half_night_upt=sheet.cell(row = match_date,column=17).value
	half_night_emp=sheet.cell(row = match_date,column=11).value

	FROM = "RPI@texmo.net"
	TO = ["ahs@texmo.net","nta@texmo.net","psu@texmo.net","svu@texmo.net","bks@texmo.net"]
#	TO = ["ahs@texmo.net"] #,"nta@texmo.net","svu@texmo.net"]
	SUBJECT = "TEXMO INDUSTRIES FURNACE UPT "+current_date
	DATA = str(half_night_unit)
	DATA1= str(half_night_charge_metal)
	DATA2= str(half_night_upt)
	DATA3= str(half_night_emp)
#	TEXT ="\n"+"DATE : "+current_date +"\n"+ "FULL NIGHT UPT DETAILES     :"+"\n"+"     UNIT CONSUMPTION : "+ DATA +"\n"+"     CHARGED METAL IN TON : " +DATA1+"\n"+"     UNITS PER TON         : "+DATA2

	TEXT ="\n"+"DATE : "+current_date +"\n"+ "HALF NIGHT UPT DETAILES:"+"\n"+"UNIT CONSUMPTION : "+ DATA +"\n"+"CHARGED METAL IN TON : " +DATA1+"\n"+"UNITS PER TON: "+DATA2+"\n"+"EMPLOYE NAME :"+DATA3


	message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

	%s

	""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	server = smtplib.SMTP(SERVER)
	server.sendmail(FROM, TO, message)
	print ("HALF NIGHT MAIL SUCCESS FULLY SENDED")
	server.quit()


def full_night_formula():
            global current_date
            print("******************FULL FORMULA**************")
            workbook = load_workbook(path_to_file)
            sheet = workbook.active
            r=sheet.max_row
            for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
                for cell in row:
                    if cell.value ==(current_date):
                        re=sheet.cell(row=cell.row, column=cell.column)
                        print("content Row: ",cell.row)
                        rr=cell.row
            match_date=rr
            print(match_date)

            sheet['L1'] = "FULL NIGHT (UNIT)"
            energy_full_night=sheet.cell(row = match_date,column=2).value
            yesterday_half_night=sheet.cell(row = match_date-1,column=4).value
            full_night_unit=round(((energy_full_night-yesterday_half_night)*1000000),2)
            print("Energy Meter Reading : "+ str(energy_full_night))
            print("Unit Consumption :" +str(full_night_unit))

#           o='=((B%d-D%d)*1000000)'% (match_date,match_date-1)
            sheet.cell(column=12, row=match_date, value=full_night_unit)
            sheet['O1'] = "FULL NIGHT (UPT)"
            full_night_charge_metal=sheet.cell(row=match_date,column=5).value
            print("charge",full_night_charge_metal)
            re=str(full_night_charge_metal)
            res=float(re)
            full_night_upt=round((full_night_unit / res),2)
            print("UPT :" +str(full_night_upt))

#            q='=(L%d/E%d)'% (match_date,match_date)
            sheet.cell(column=15, row=match_date, value=full_night_upt)
            workbook.save(path_to_file)

def day_shift_formula():
            global current_date
            print("******************DAY FORMULA**************")
            workbook = load_workbook(path_to_file)
            sheet = workbook.active
            r=sheet.max_row
            for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
                for cell in row:
                    if cell.value ==(current_date):
                        re=sheet.cell(row=cell.row, column=cell.column)
                        print("content Row: ",cell.row)
                        rr=cell.row
            match_date=rr
            print(match_date)
            sheet['M1'] = "DAY SHIFT (UNIT)"
            energy_day_shift=sheet.cell(row = match_date,column=3).value
            today_full_night=sheet.cell(row = match_date,column=2).value
            day_shift_unit=round(((energy_day_shift-today_full_night)*1000000),2)
            print("Energy Meter Reading : "+ str(energy_day_shift))
            print("Unit Consumption :" +str(day_shift_unit))
#            o='=((C%d-B%d)*1000000)'% (match_date,match_date)

            sheet.cell(column=13, row=match_date, value=day_shift_unit)

            sheet['P1'] = "DAY SHIFT (UPT)"
            #DAY Shift Charged Metal
            day_shift_charge_metal=sheet.cell(row=match_date,column=6).value
            print("DAY SHift Charged MEtal",day_shift_charge_metal)
            re=str(day_shift_charge_metal)
            res=float(re)
            #DAY Shift UPT
            day_shift_upt=round((day_shift_unit / res),2)
            print("DAY SHIFT UPT :" +str(day_shift_upt))
#            q='=(M%d/F%d)'% (match_date,match_date)

            sheet.cell(column=16, row=match_date, value=day_shift_upt)
            workbook.save(path_to_file)


def half_night_formula():
            global current_date
            print("******************HALF FORMULA**************")
            workbook = load_workbook(path_to_file)
            sheet = workbook.active
            r=sheet.max_row
            for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
                for cell in row:
                    if cell.value ==(current_date):
                        re=sheet.cell(row=cell.row, column=cell.column)
                        print("content Row: ",cell.row)
                        rr=cell.row
            match_date=rr
            print(match_date)
            sheet['N1'] = "HALF NIGHT (UNIT)"
            energy_half_night=sheet.cell(row = match_date,column=4).value
            today_day_shift=sheet.cell(row = match_date,column=3).value
            half_night_unit=round(((energy_half_night-today_day_shift)*1000000),2)
            print("Energy Meter Reading : "+ str(energy_half_night))
            print("Unit Consumption :" +str(half_night_unit))

#            o='=((D%d-C%d)*1000000)'% (match_date,match_date)
            sheet.cell(column=14, row=match_date, value=half_night_unit)

            sheet['Q1'] = "HALF NIGHT (UPT)"
            half_night_charge_metal=sheet.cell(row=match_date,column=7).value
            print("charge",half_night_charge_metal)
            re=str(half_night_charge_metal)
            res=float(re)
            half_night_upt=round((half_night_unit / res),2)
            print("UPT :" +str(half_night_upt))

#            q='=(N%d/G%d)'% (match_date,match_date)
            sheet.cell(column=17, row=match_date, value=half_night_upt)

            workbook.save(path_to_file)


@app.route("/field_data", methods=['GET', 'POST'])
def field_data():
    global current_date
    workbook = load_workbook(path_to_file)
    sheet = workbook.active
    if request.method == 'POST':

        date_get_from=request.form['date']
        get_date=date_get_from
        date = get_date
        convert_date=str(date)
        unorder_date = convert_date
        create_list = unorder_date.split("-")
        create_list.reverse()
        order_list=create_list
        list_to_str = ""
        for i in order_list:
          list_to_str += i
        order_str=list_to_str
        format_date=order_str
        format_date = format_date[:2] + '-' + format_date[2:]
        final_format_date = format_date
        final_format_date = final_format_date[:5] + '-' + final_format_date[5:]
        print(final_format_date)
        shift_get_from=request.form['shift']
        shift=shift_get_from
        lmt_get_from=request.form['lmt']
        lmt=lmt_get_from
        emp_get_from=request.form['emp']
        emp=emp_get_from
        current_date=final_format_date
        r=sheet.max_row
        for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
            for cell in row:
                if cell.value ==(current_date):
                    re=sheet.cell(row=cell.row, column=cell.column)
                    print("content Row: ",cell.row)
                    rr=cell.row
        match_date=rr
        print(match_date)
        sheet['E1'] = "CHARGE WEIGHT (12:30AM to 7:30 AM)"
        sheet['F1'] = "CHARGE WEIGHT (7:30AM to 4.00 PM)"
        sheet['G1'] = "CHARGE WEIGHT (4.00PM to 12.30 AM)"
        sheet['I1'] = "EMPLOYE NAME FULL NIGHT"
        sheet['J1'] = "EMPLOYE NAME DAY"
        sheet['K1'] = "EMPLOYE NAME HALF NIGHT"

        if shift==("FULL NIGHT"):
            sheet.column_dimensions['E'].width = 30
            sheet.cell(column=5, row=match_date, value=lmt)
            sheet.column_dimensions['I'].width = 30
            sheet.cell(column=9, row=match_date, value=emp)
            workbook.save(path_to_file)
            full_night_formula()

        elif shift==("DAY"):
            sheet.column_dimensions['F'].width = 30
            sheet.cell(column=6, row=match_date, value=lmt)
            sheet.column_dimensions['J'].width = 30
            sheet.cell(column=10, row=match_date, value=emp)
            workbook.save(path_to_file)
            day_shift_formula()

        elif shift==("HALF NIGHT"):
            sheet.column_dimensions['E'].width = 30
            sheet.cell(column=7, row=match_date, value=lmt)
            sheet.column_dimensions['K'].width = 30
            sheet.cell(column=11, row=match_date, value=emp)
            workbook.save(path_to_file)
            half_night_formula()

        templateData = {
            'date' :final_format_date,
            'shift' : shift,
            'lmt' :lmt,
            'emp' : emp
             }

        return render_template('disply.html',**templateData)

#   ------------------------------------------------------------------
@app.route("/admin_entry", methods=['GET', 'POST'])
def admin_entry():
    global current_date
    workbook = load_workbook(path_to_file)
    sheet = workbook.active
    if request.method == 'POST':
        date_get_from=request.form['date']
        get_date=date_get_from
        date = get_date
        convert_date=str(date)
        unorder_date = convert_date
        create_list = unorder_date.split("-")
        create_list.reverse()
        order_list=create_list
        list_to_str = ""
        for i in order_list:
          list_to_str += i
        order_str=list_to_str
        format_date=order_str
        format_date = format_date[:2] + '-' + format_date[2:]
        final_format_date = format_date
        final_format_date = final_format_date[:5] + '-' + final_format_date[5:]
        print(final_format_date)
        shift_get_from=request.form['shift']
        shift=shift_get_from
        lmt_get_from=request.form['lmt']
        lmt=lmt_get_from
        emp_get_from=request.form['emp']
        emp=emp_get_from
        energy_get_from=request.form['energy']
        energy=energy_get_from
        current_date=final_format_date
        r=sheet.max_row
        for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
            for cell in row:
                if cell.value ==(current_date):
                    re=sheet.cell(row=cell.row, column=cell.column)
                    print("content Row: ",cell.row)
                    rr=cell.row
        match_date=rr
        print(match_date)
        sheet['E1'] = "CHARGE WEIGHT (12:30AM to 7:30 AM)"
        sheet['F1'] = "CHARGE WEIGHT (7:30AM to 4.00 PM)"
        sheet['G1'] = "CHARGE WEIGHT (4.00PM to 12.30 AM)"
        sheet['I1'] = "EMPLOYE NAME FULL NIGHT"
        sheet['J1'] = "EMPLOYE NAME DAY"
        sheet['K1'] = "EMPLOYE NAME HALF NIGHT"

        if shift==("FULL NIGHT"):
            sheet.column_dimensions['E'].width = 30
            sheet.cell(column=5, row=match_date, value=lmt)
            sheet.column_dimensions['I'].width = 30
            sheet.cell(column=9, row=match_date, value=emp)
            sheet.cell(column=2, row=match_date, value=energy)
            workbook.save(path_to_file)
            full_night_formula()

        elif shift==("DAY"):
            sheet.column_dimensions['F'].width = 30
            sheet.cell(column=6, row=match_date, value=lmt)
            sheet.column_dimensions['J'].width = 30
            sheet.cell(column=10, row=match_date, value=emp)
            sheet.cell(column=3, row=match_date, value=energy)
            workbook.save(path_to_file)
            day_shift_formula()

        elif shift==("HALF NIGHT"):
            sheet.column_dimensions['E'].width = 30
            sheet.cell(column=7, row=match_date, value=lmt)
            sheet.column_dimensions['K'].width = 30
            sheet.cell(column=11, row=match_date, value=emp)
            sheet.cell(column=4, row=match_date, value=energy)
            workbook.save(path_to_file)
            half_night_formula()

        templateData = {
            'date' :final_format_date,
            'shift' : shift,
            'lmt' :lmt,
            'emp' : emp,
            'energy' : energy
             }

        return render_template('disply.html',**templateData)
#   ------------------------------------------------------------------

@app.route("/field_data_user", methods=['GET', 'POST'])
def field_data_user():
    global current_date
    now = datetime.now()
    workbook = load_workbook(path_to_file)
    sheet = workbook.active
    if request.method == 'POST':
        lmt_get_from=request.form['lmt']
        lmt=lmt_get_from
        emp_get_from=request.form['emp']
        emp=emp_get_from
        current_time=now.strftime("%I:%M:%S:%P")
        print("TIME : " + current_time)
        current_date=now.strftime("%d-%m-%Y")
        print("DATE : " + current_date)
        h=now.strftime("%H")
        hour=str(h)
        print(hour)
        sheet['E1'] = "CHARGE WEIGHT (12:30AM to 7:30 AM)"
        sheet['F1'] = "CHARGE WEIGHT (7:30AM to 4.00 PM)"
        sheet['G1'] = "CHARGE WEIGHT (4.00PM to 12.30 AM)"
        sheet['I1'] = "EMPLOYE NAME FULL NIGHT"
        sheet['J1'] = "EMPLOYE NAME DAY"
        sheet['K1'] = "EMPLOYE NAME HALF NIGHT"
        half=["00","01","02","03","04","05","06"]
        full=["07","08","09","10","11","12","13","14","15"]
        day=["16","17","18","19","20","21","22","23"]
        r=sheet.max_row
        for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1 , values_only=False):
            for cell in row:
                if cell.value ==(current_date):
                    re=sheet.cell(row=cell.row, column=cell.column)
                    print("content Row: ",cell.row)
                    rr=cell.row
        match_date=rr
        print(match_date)
        if hour in full:
            shift="FULL NIGHT"
            sheet.column_dimensions['E'].width = 30
            sheet.cell(column=5, row=match_date, value=lmt)
            sheet.column_dimensions['I'].width = 30
            sheet.cell(column=9, row=match_date, value=emp)
            workbook.save(path_to_file)
            full_night_formula()
            full_night_mail()

        elif hour in day:
            shift="DAY"
            sheet.column_dimensions['F'].width = 30
            sheet.cell(column=6, row=match_date, value=lmt)
            sheet.column_dimensions['J'].width = 30
            sheet.cell(column=10, row=match_date, value=emp)
            workbook.save(path_to_file)
            day_shift_formula()
            day_shift_mail()

        elif hour in half:
            shift="HALF NIGHT"
            sheet.column_dimensions['E'].width = 30
            sheet.cell(column=7, row=match_date, value=lmt)
            sheet.column_dimensions['K'].width = 30
            sheet.cell(column=11, row=match_date, value=emp)
            workbook.save(path_to_file)
            half_night_formula()
            half_night_mail()

        templateData = {
            'lmt' :lmt,
            'emp' : emp,
            'date': current_time,
            'shift':shift
             }
        return render_template('disply.html',**templateData)

#-------------Energy Meter Data View Thread------------------------
@app.route("/energy_data_update")
def energy_data_update():
        global kwh_data
        global unit
        data = client.read_holding_registers(HMI_Doller_Address , 2 , slave = 1)
        if not data.isError():
            decoder = BinaryPayloadDecoder.fromRegisters(data.registers,  Endian.Big, wordorder=Endian.Little)
            address_result   = decoder.decode_32bit_float()
            string_convert   = str(address_result)
            length_of_number = len(string_convert)
            print ("The Register Address Value Is :  ",(address_result))
            print("Total Length Of Character : ",(length_of_number))
            if (length_of_number==9):
                    Round_Value = round(address_result)
                    kwh = Round_Value
                    kwh_data=kwh
                    print (kwh_data)
            elif (length_of_number==10):
                    Round_Value = round(address_result)
                    kwh = Round_Value/1000
                    kwh_data=round(kwh,14)
                    print ("KWh",kwh)
                    print (kwh_data)
            elif (length_of_number==11):
                    Round_Value = round(address_result)
                    kwh = Round_Value/100000
                    kwh_data=round(kwh,14)
                    print ("MWh",kwh_data)
                    print (kwh_data)
            elif (length_of_number==12):
                    Round_Value = round(address_result)
                    kwh = Round_Value/10000000
                    kwh_data=round(kwh,14)
                    print ("MWh",kwh_data)
                    print (kwh_data)
            elif (length_of_number==13):
                    Round_Value = address_result
                    kwh = Round_Value/1000000000
                    kwh_data=round(kwh,14)
                    unit="GWh"
                    print ("MWh",kwh_data)
                    print (kwh_data)
            elif (length_of_number==3):
                    Round_Value = address_result
                    kwh = Round_Value/1000000000
                    kwh_data=round(kwh,14)
                    unit="GWh"
                    print ("MWh",kwh_data)
                    print (kwh_data)


            else:
                    print ("Something Wrong If You read Energy meter data shuold be contain minimum 10 integers \n Please Check KWh Or MWh or GWh")
        else :
             print ("ERROR")

        return jsonify(kwh_data=kwh_data,unit=unit)

@app.route("/asd")
def asd():
        now = datetime.now()
        current_time=now.strftime("%I:%M:%S:%P")
        current_date=now.strftime("%d-%m-%Y")
        rpi_temp = CPUTemperature()
        cpu=(rpi_temp.temperature)
        return jsonify(cpu=cpu,date=current_date,time=current_time)

@app.route("/view")
def view():
	workbook = load_workbook(path_to_file)
	sheet = workbook.active
#	sheet = pd.read_excel(path_to_file)
	return render_template("view.html", sheet=sheet)

if __name__ == "__main__":
    ip_address ="172.16.4.237"
    HMI_Doller_Address = 746
    client = ModbusTcpClient(ip_address,port = 502)
    client.connect()
    app.run(host='172.16.3.119',port =5001 ,debug=True)


#if program already running you type this command line
#sudo ps aux | grep -i main.py
#find the Process_ID
#sudo kill -9 process_ID
