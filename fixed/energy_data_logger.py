#!/bin/bash
#!/bin/sh

#!/usr/bin/python3
import os
from pymodbus.constants import Endian
from pymodbus.payload import BinaryPayloadDecoder
from pymodbus.client import ModbusSerialClient
from pymodbus.client import ModbusTcpClient
from fractions import Fraction
import time
from openpyxl import *
import os.path
from pathlib import Path
from datetime import datetime

path_to_file =("Furnace_UPT/rpiWebServer/Furnace_UPT.xlsx")
path = Path(path_to_file)
access=False

if path.is_file():
    print ("The file ")
   
    
else:
    workbook = Workbook()
    sheet = workbook.active
#    sheet.protection.sheet = True
#    sheet.protection.password = '433'
#    sheet.protection.enable()
    
    

    workbook.save("Furnace_UPT/rpiWebServer/Furnace_UPT.xlsx")
    print("The File Not In This Directory now The File Created")
    


def full_night():

    workbook=load_workbook(path_to_file)
    sheet=workbook.active
    sheet['A1'] = "DATE"
    sheet['B1'] = "FULL NIGHT (7:30 AM)"
        
    now = datetime.now()
    current_time=now.strftime("%I:%M:%S:%P")
    print("TIME : " + current_time)
    current_date=now.strftime("%d-%m-%Y")
    print("DATE : " + current_date)
    
    ip_address ="172.16.4.237"
    HMI_Doller_Address = 746
    client = ModbusTcpClient(ip_address,port = 502)
    client.connect()
    data = client.read_holding_registers(HMI_Doller_Address , 2 , slave = 1)
    decoder = BinaryPayloadDecoder.fromRegisters(data.registers,  Endian.Big, wordorder=Endian.Little)
    address_result   = decoder.decode_32bit_float()
    string_convert   = str(address_result)
    length_of_number = len(string_convert)
    print ("The Register Address Value Is :  ",(address_result))
    Round_Value = address_result
    kwh = Round_Value/1000000000
    kwh_data=round(kwh,14)
    print("FULL" , kwh_data)
    data1=kwh_data

    i = 1
    for col in sheet['B']:
        if col.value is not None:
            i += 1
    print('First row empty: ',i) 

    new_line=i
    print("B",new_line)
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.cell(column=1, row=new_line, value= current_date)
    sheet.cell(column=2, row=new_line, value=data1)

   
       
    
    workbook.save(path_to_file)
    
def day_shift():
    workbook=load_workbook(path_to_file)
    sheet=workbook.active
    sheet['C1'] = "DAY SHIFT (4:00 PM)"
#    sheet['M1'] = "DAY SHIFT (UNIT)"
    ip_address ="172.16.4.237"
    HMI_Doller_Address = 746
    client = ModbusTcpClient(ip_address,port = 502)
    client.connect()
    data = client.read_holding_registers(HMI_Doller_Address , 2 , slave = 1)
    decoder = BinaryPayloadDecoder.fromRegisters(data.registers,  Endian.Big, wordorder=Endian.Little)
    address_result   = decoder.decode_32bit_float()
    string_convert   = str(address_result)
    length_of_number = len(string_convert)
    print ("The Register Address Value Is :  ",(address_result))
    Round_Value = address_result
    kwh = Round_Value/1000000000
    kwh_data=round(kwh,14)
    print("DAY" , kwh_data)
    data1=kwh_data
    
    i = 1
    for col in sheet['C']:
        if col.value is not None:
            i += 1
    print('First row empty: ',i) 
  
    new_line=i
    print("C",new_line)
    sheet.column_dimensions['C'].width = 20
    sheet.cell(column=3, row=new_line, value=data1)
    workbook.save(path_to_file)
    
def half_night():
    workbook=load_workbook(path_to_file)
    sheet=workbook.active
    sheet['D1'] = "HALF NIGHT (12.30 AM)"
 #   sheet['N1'] = "HLAF NIGHT (UNIT)"
    ip_address ="172.16.4.237"
    HMI_Doller_Address = 746
    client = ModbusTcpClient(ip_address,port = 502)
    client.connect()
    data = client.read_holding_registers(HMI_Doller_Address , 2 , slave = 1)
    decoder = BinaryPayloadDecoder.fromRegisters(data.registers,  Endian.Big, wordorder=Endian.Little)
    address_result   = decoder.decode_32bit_float()
    string_convert   = str(address_result)
    length_of_number = len(string_convert)
    print ("The Register Address Value Is :  ",(address_result))
    Round_Value = address_result
    kwh = Round_Value/1000000000
    kwh_data=round(kwh,14)
    print("HALF" , kwh_data)
    data1=kwh_data
    
    i = 1
    for col in sheet['D']:
        if col.value is not None:
            i += 1
    print('First row empty: ',i) 
  
    new_line=i
    print("D",new_line)
    sheet.column_dimensions['D'].width = 20

    sheet.cell(column=4, row=new_line, value=data1)
    workbook.save(path_to_file)

def reboot_full():
	os.system("sudo reboot")
    
while True:          
	#        FULL= "07:20:00:am"
	#        DAY="04:00:00:pm"
	#        HALF="12:30:00:am"
	FULL= "07:30:00:am"
	DAY="04:00:00:pm"
	HALF="12:30:00:am"

	re_full="03:00:00:am"
	re_day="09:00:00:am"
	re_half="12:00:00:pm"
	re_a="03:00:00:pm"
	re_b="07:00:00:pm"
	re_c="10:00:00:pm"
	re_d="06:00:00:am"
	re_e="12:00:00:am"
	re_f="02:00:00:am"

	now = datetime.now()
	current_time=now.strftime("%I:%M:%S:%P")
	print("TIME : " + current_time)
	current_date=now.strftime("%d-%m-%Y")
	print("DATE : " + current_date)
	if (current_time==FULL):
		full_night()
	elif (current_time==DAY):
		day_shift()          
	elif (current_time==HALF):
		half_night()
	elif (current_time==re_full):
		reboot_full()
	elif (current_time==re_day):
		reboot_full()
	elif (current_time==re_half):
		reboot_full()
	elif (current_time==re_a):
		reboot_full()
	elif (current_time==re_b):
		reboot_full()
	elif (current_time==re_c):
		reboot_full()
	elif (current_time==re_d):
		reboot_full()
	elif (current_time==re_e):
		reboot_full()
	elif (current_time==re_f):
		reboot_full()

	time.sleep(1)
