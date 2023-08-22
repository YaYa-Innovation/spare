#!/bin/bash
#!/bin/sh
#!/usr/bin/python3

from openpyxl import *
from datetime import datetime
import time
import smtplib

SERVER = "172.16.0.8"

path_to_file =("Furnace_UPT/rpiWebServer/Furnace_UPT.xlsx")

def full_night_mail():
	workbook = load_workbook(path_to_file,data_only=True)
	sheet = workbook.active
	now = datetime.now()
	current_time=now.strftime("%I:%M:%S:%P")
	print("TIME : " + current_time)
	current_date=now.strftime("%d-%m-%Y")
	print("DATE : " + current_date)
	r=sheet.max_row
	for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
	        for cell in row:
	                if cell.value ==(current_date):
	                        re=sheet.cell(row=cell.row, column=cell.column)
	                        print("content Row: ",cell.row)
	                        rr=cell.row
	match_date=rr
	print(match_date)
	energy_full_night=sheet.cell(row = match_date,column=2).value
	yesterday_half_night=sheet.cell(row = match_date-1,column=4).value
	result=round(((energy_full_night-yesterday_half_night)*1000000),2)

	charge_metal=sheet.cell(row=match_date,column=6).value
	print("charge",charge_metal)
	re=str(charge_metal)
	res=float(re)

	upt=round((result / res),2)

	print("Energy Meter Reading : "+ str(energy_full_night))
	print("Unit Consumption :" +str(result))
	print("UPT :" +str(upt))

	FROM = "RPI@texmo.net"
	TO = ["ahs@texmo.net","nta@texmo.net","svu@texmo.net"]
	SUBJECT = "TEXMO INDUSTRIES FURNACE UPT "+current_date
	DATA = str(result)
	DATA1= str(charge_metal)
	DATA2= str(upt)

	TEXT ="\n"+"DATE : "+current_date +"\n"+ "FULL NIGHT UPT DETAILES :"+"\n"+"UNIT CONSUMPTION : "+ DATA +"\n"+"CHARGED METAL IN TON : " +DATA1+"\n"+"UNITS PER TON : "+DATA2

	message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

	%s

	""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	server = smtplib.SMTP(SERVER)
	server.sendmail(FROM, TO, message)
	print ("FULL NIGHT MAIL SUCCESS FULLY SENDED")
	server.quit()

def day_shift_mail():
	workbook = load_workbook(path_to_file,data_only=True)
	sheet = workbook.active
	now = datetime.now()
	current_time=now.strftime("%I:%M:%S:%P")
	print("TIME : " + current_time)
	current_date=now.strftime("%d-%m-%Y")
	print("DATE : " + current_date)
	r=sheet.max_row
	for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
	        for cell in row:
	                if cell.value ==(current_date):
	                        re=sheet.cell(row=cell.row, column=cell.column)
	                        print("content Row: ",cell.row)
	                        rr=cell.row
	match_date=rr
	print(match_date)


	energy_full_night=sheet.cell(row = match_date,column=3).value
	yesterday_half_night=sheet.cell(row = match_date,column=2).value
	result=round(((energy_full_night-yesterday_half_night)*1000000),2)

	charge_metal=sheet.cell(row=match_date,column=5).value
	print("charge",charge_metal)
	re=str(charge_metal)
	res=float(re)

	upt=round((result / res),2)

	print("Energy Meter Reading : "+ str(energy_full_night))
	print("Unit Consumption :" +str(result))
	print("UPT :" +str(upt))

	FROM = "RPI@texmo.net"
	TO = ["ahs@texmo.net","nta@texmo.net","svu@texmo.net"]
	SUBJECT = "TEXMO INDUSTRIES FURNACE UPT "+current_date
	DATA = str(result)
	DATA1= str(charge_metal)
	DATA2= str(upt)

	TEXT ="\n"+"DATE : "+current_date +"\n"+ "DAY SHIFT UPT DETAILES :"+"\n"+"UNIT CONSUMPTION : "+ DATA +"\n"+"CHARGED METAL IN TON : " +DATA1+"\n"+"UNITS PER TON : "+DATA2


	message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

	%s

	""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	server = smtplib.SMTP(SERVER)
	server.sendmail(FROM, TO, message)
	print ("DAY SHIFT MAIL SUCCESS FULLY SENDED")
	server.quit()

def half_night_mail():
	workbook = load_workbook(path_to_file,data_only=True)
	sheet = workbook.active
	now = datetime.now()
	current_time=now.strftime("%I:%M:%S:%P")
	print("TIME : " + current_time)
	current_date=now.strftime("%d-%m-%Y")
	print("DATE : " + current_date)
	r=sheet.max_row
	for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
	        for cell in row:
	                if cell.value ==(current_date):
	                        re=sheet.cell(row=cell.row, column=cell.column)
	                        print("content Row: ",cell.row)
	                        rr=cell.row
	match_date=rr
	print(match_date)

	energy_full_night=sheet.cell(row = match_date,column=4).value
	yesterday_half_night=sheet.cell(row = match_date,column=3).value
	result=round(((energy_full_night-yesterday_half_night)*1000000),2)

	charge_metal=sheet.cell(row=match_date,column=7).value
	print("charge",charge_metal)
	re=str(charge_metal)
	res=float(re)

	upt=round((result / res),2)

	print("Energy Meter Reading : "+ str(energy_full_night))
	print("Unit Consumption :" +str(result))
	print("UPT :" +str(upt))

	FROM = "RPI@texmo.net"
	TO = ["ahs@texmo.net","nta@texmo.net","svu@texmo.net"]
	SUBJECT = "TEXMO INDUSTRIES FURNACE UPT "+current_date
	DATA = str(result)
	DATA1= str(charge_metal)
	DATA2= str(upt)

	TEXT ="\n"+"DATE : "+current_date +"\n"+ "HALF NIGHT UPT DETAILES :"+"\n"+"UNIT CONSUMPTION : "+ DATA +"\n"+"CHARGED METAL IN TON : " +DATA1+"\n"+"UNITS PER TON : "+DATA2


	message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

	%s

	""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	server = smtplib.SMTP(SERVER)
	server.sendmail(FROM, TO, message)
	print ("HALF NIGHT MAIL SUCCESS FULLY SENDED")
	server.quit()

while True:
	print ("MAIL Timer Start")
	FULL= "07:30:00:am"
	DAY="04:00:00:pm"
	HALF="12:30:40:am"
	now = datetime.now()
	current_time=now.strftime("%I:%M:%S:%P")
	print("TIME : " + current_time)
	current_date=now.strftime("%d-%m-%Y")
	print("DATE : " + current_date)
	if (current_time==FULL):
	    full_night_mail()
	elif (current_time==DAY):
	    day_shift_mail()          
	elif (current_time==HALF):
	    half_night_mail()
	time.sleep(1)

